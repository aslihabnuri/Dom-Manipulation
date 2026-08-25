# -*- coding: utf-8 -*-
"""Daftar Affiliate TikTok Toni Black, versi 2: 8 file Kalodata + kolom Quantity Terjual."""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from data_batch1 import agregasi, PERIODE, SUMBER, ROWS

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "v8.xlsx")

FONT = "Arial"
NAVY = "14213D"
HDR = "1F4E79"
HDR_INPUT = "7F6000"
WHITE = "FFFFFF"
GREY = "E8E8E8"
YELLOW = "FFF2CC"
ST = (("Belum dihubungi", "FCE4E4", "9C0006"),
      ("Sedang dihubungi", "FFF2CC", "9C6500"),
      ("Kirim Sampel", "DFF3E0", "006100"))

ROW_HDR = 5
ROW_START = 6
SPARE = 1

FMT_RP = '"Rp" #,##0'
FMT_NUM = '#,##0'

thin = Side(style="thin", color="C9CFD8")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

data = agregasi()
ROW_LAST = ROW_START + len(data) - 1
ROW_END = ROW_LAST + SPARE

wb = Workbook()
ws = wb.active
ws.title = "Daftar Affiliate"
ws.sheet_properties.tabColor = "C00000"
ws.sheet_format.defaultRowHeight = 18
ws.sheet_format.customHeight = True

ws["A1"] = "DAFTAR AFFILIATE TIKTOK - TONI BLACK"
ws["A1"].font = Font(name=FONT, size=16, bold=True, color=WHITE)
ws["A1"].alignment = Alignment(vertical="center", indent=1)
ws.merge_cells("A1:N1")
for c in ws["A1:N1"][0]:
    c.fill = PatternFill("solid", fgColor=NAVY)
ws.row_dimensions[1].height = 32

ws["A2"] = ("{} file export Kalodata, periode {}. Total {} kreator, diurutkan dari GMV terbesar.").format(
    len(SUMBER), PERIODE, len(data))
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
ws["A2"].alignment = Alignment(vertical="center", indent=1)
ws.merge_cells("A2:N2")

ws["A3"] = ("Cara pakai: 5 kolom terakhir (judul kuning) diisi manual. Dihubungi Via Apa dan Progress memakai dropdown, "
            "klik selnya lalu pilih. Kolom No berisi rumus otomatis, jangan diketik.")
ws["A3"].font = Font(name=FONT, size=9, color="806000")
ws["A3"].fill = PatternFill("solid", fgColor=YELLOW)
ws["A3"].alignment = Alignment(vertical="center", wrap_text=True, indent=1)
ws.merge_cells("A3:N3")
ws.row_dimensions[3].height = 30
ws.row_dimensions[4].height = 6

HEADERS = ["No", "Nama Kreator", "Handle Kreator", "Followers", "GMV (Penjualan)", "Quantity Terjual",
           "Harga Jual Rata-Rata", "Jumlah Video", "Jumlah Live", "Nomor Whatsapp",
           "Product yang Dikirim", "Dihubungi Via Apa?", "Progress", "Komisi"]
INPUT_COLS = (10, 11, 12, 13, 14)

for i, h in enumerate(HEADERS):
    col = 1 + i
    c = ws.cell(row=ROW_HDR, column=col, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=HDR_INPUT if col in INPUT_COLS else HDR)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
ws.row_dimensions[ROW_HDR].height = 34

# Kolom data (A sampai I) dan Progress (M) diberi gaya. Kolom input lain dibiarkan
# polos supaya ukuran file tetap muat pada batas unggah Google Drive.
for idx in range(ROW_START, ROW_END + 1):
    i = idx - ROW_START
    isi = data[i] if i < len(data) else None

    a = ws.cell(row=idx, column=1, value='=IF($B{0}<>"",ROW()-{1},"")'.format(idx, ROW_HDR))
    a.font = Font(name=FONT, size=10, bold=True, color="595959")
    a.fill = PatternFill("solid", fgColor=GREY)
    a.alignment = Alignment(horizontal="center", vertical="center")

    if isi:
        ws.cell(row=idx, column=2, value=isi["nama"])
        ws.cell(row=idx, column=3, value=isi["handle"])
        ws.cell(row=idx, column=4, value=isi["followers"])
        ws.cell(row=idx, column=5, value=round(isi["gmv"]))
        ws.cell(row=idx, column=6, value=isi["qty"])
        ws.cell(row=idx, column=7, value=round(isi["aov"]))
        ws.cell(row=idx, column=8, value=isi["video"])
        ws.cell(row=idx, column=9, value=isi["live"])
        ws.cell(row=idx, column=13, value="Belum dihubungi")

    for col in list(range(2, 10)) + [13]:
        c = ws.cell(row=idx, column=col)
        c.font = Font(name=FONT, size=10)
        if col in (4, 6, 8, 9):
            c.number_format = FMT_NUM
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        elif col in (5, 7):
            c.number_format = FMT_RP
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        elif col == 13:
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(vertical="center", indent=1)

# format kolom input lewat satu sel contoh di baris pertama supaya format tetap terbawa
ws.cell(row=ROW_START, column=10).number_format = "@"
ws.cell(row=ROW_START, column=14).number_format = FMT_RP

rng = "{0}{1}:{0}{2}"
dv_via = DataValidation(type="list", formula1='"DM,Whatsapp"', allow_blank=True, showDropDown=False)
dv_via.promptTitle, dv_via.prompt = "Dihubungi Via Apa?", "Pilih DM atau Whatsapp."
ws.add_data_validation(dv_via)
dv_via.add(rng.format("L", ROW_START, ROW_END))

dv_prog = DataValidation(type="list", formula1='"Belum dihubungi,Sedang dihubungi,Kirim Sampel"',
                         allow_blank=True, showDropDown=False)
dv_prog.promptTitle, dv_prog.prompt = "Progress", "Pilih status terkini kreator ini."
ws.add_data_validation(dv_prog)
dv_prog.add(rng.format("M", ROW_START, ROW_END))

dv_num = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv_num.errorTitle, dv_num.error = "Nilai tidak valid", "Komisi harus angka dan tidak boleh negatif."
ws.add_data_validation(dv_num)
dv_num.add(rng.format("N", ROW_START, ROW_END))

for teks, warna, warna_font in ST:
    ws.conditional_formatting.add(rng.format("M", ROW_START, ROW_END), CellIsRule(
        operator="equal", formula=['"{0}"'.format(teks)],
        fill=PatternFill("solid", start_color=warna, end_color=warna),
        font=Font(name=FONT, size=10, bold=True, color=warna_font)))

ws.auto_filter.ref = "A{0}:N{1}".format(ROW_HDR, ROW_END)
ws.freeze_panes = "D{0}".format(ROW_START)

for col, w in zip("ABCDEFGHIJKLMN",
                  (5, 26, 23, 12, 18, 15, 18, 12, 11, 18, 24, 18, 18, 15)):
    ws.column_dimensions[col].width = w

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = "{0}:{0}".format(ROW_HDR)


wb.save(OUT)
print("saved | kreator:", len(data))
