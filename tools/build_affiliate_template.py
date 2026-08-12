#!/usr/bin/env python3
"""Build the Toni Black affiliate tracker workbook.

Four sheets: a Dashboard that summarises itself, the affiliate list, a twelve
month GMV log, and a Referensi sheet holding the dropdown values and the tier
thresholds so nothing is hardcoded inside a formula.

The point of the list is not to rank affiliates by follower count. Follower
count is what people pick on and it predicts sales badly, so the sheet computes
GMV per 1.000 followers and GMV per 1.000 views alongside it. A micro affiliate
who sells more per follower than a macro one is the finding worth having, and it
is invisible in a list sorted by size.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

OUT = "/home/user/Dom-Manipulation/content/affiliate/Toni_Black_Affiliate_Tracker.xlsx"

INK, PAPER = "282828", "FFFFFF"          # brand black / clean white
GREY, STEEL = "4F5052", "CCCCCC"
INPUT_FILL = "FFFDF0"                    # pale wash marking a column you type into
CALC_FILL = "F4F4F4"
HELP_FILL = "EDEDED"

F = "Arial"
H_FONT = Font(name=F, size=10, bold=True, color=PAPER)
LBL = Font(name=F, size=10, bold=True, color=INK)
IN_FONT = Font(name=F, size=10, color="0000FF")     # blue: you type here
CA_FONT = Font(name=F, size=10, color="000000")     # black: formula
NOTE = Font(name=F, size=9, italic=True, color=GREY)
TITLE = Font(name=F, size=16, bold=True, color=INK)
SUB = Font(name=F, size=10, color=GREY)

thin = Side(style="thin", color=STEEL)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
RP = '"Rp"#,##0;("Rp"#,##0);-'
NUM = '#,##0;(#,##0);-'
PCT = '0.0%;(0.0%);-'

FIRST, LAST = 6, 105                     # data rows on Daftar Affiliate

wb = Workbook()

# ─────────────────────────────── REFERENSI ──────────────────────────────────
# Written first so every other sheet can point at it. Thresholds live in cells,
# never inside an IF chain, so re-tiering later is an edit not a rewrite.
ref = wb.active
ref.title = "Referensi"
ref["A1"] = "REFERENSI"
ref["A1"].font = TITLE
ref["A2"] = "Ubah angka di sini kalau definisinya berubah. Semua sheet lain mengikuti."
ref["A2"].font = SUB

ref["A4"] = "TIER BERDASARKAN FOLLOWERS"
ref["A4"].font = LBL
for c, h in zip("AB", ["Batas Bawah Followers", "Tier"]):
    ref[f"{c}5"] = h
    ref[f"{c}5"].font = H_FONT
    ref[f"{c}5"].fill = PatternFill("solid", fgColor=INK)
TIERS = [(0, "Nano"), (10_000, "Micro"), (100_000, "Mid"),
         (500_000, "Macro"), (1_000_000, "Mega")]
for i, (lo, name) in enumerate(TIERS):
    ref.cell(6 + i, 1, lo).font = IN_FONT
    ref.cell(6 + i, 1).number_format = NUM
    ref.cell(6 + i, 2, name).font = CA_FONT
    for col in (1, 2):
        ref.cell(6 + i, col).border = BOX
        ref.cell(6 + i, col).fill = PatternFill("solid", fgColor=INPUT_FILL)
ref["A12"] = "Ambang ini standar industri KOL, bukan angka Toni Black. Ganti kalau perlu."
ref["A12"].font = NOTE

ref["D4"] = "PLATFORM"
ref["D4"].font = LBL
PLATFORMS = ["TikTok", "TikTok Shop", "Instagram", "Shopee Live", "YouTube"]
for i, p in enumerate(PLATFORMS):
    ref.cell(5 + i, 4, p).font = CA_FONT
    ref.cell(5 + i, 4).border = BOX

ref["F4"] = "STATUS"
ref["F4"].font = LBL
STATUSES = ["Aktif", "Baru", "Pantau", "Berhenti"]
for i, s in enumerate(STATUSES):
    ref.cell(5 + i, 6, s).font = CA_FONT
    ref.cell(5 + i, 6).border = BOX

ref["A15"] = "ARTI WARNA"
ref["A15"].font = LBL
LEGEND = [
    ("Teks biru, latar krem", "Kolom isian. Hanya ini yang Anda ketik."),
    ("Teks hitam, latar abu", "Kolom rumus. Jangan diketik, nanti rumusnya hilang."),
    ("Latar abu tua", "Kolom bantu untuk Dashboard. Boleh disembunyikan."),
]
for i, (a, b) in enumerate(LEGEND):
    ref.cell(16 + i, 1, a).font = Font(name=F, size=10, bold=True, color=INK)
    ref.cell(16 + i, 2, b).font = Font(name=F, size=10, color=INK)
ref.cell(16, 1).fill = PatternFill("solid", fgColor=INPUT_FILL)
ref.cell(16, 1).font = IN_FONT
ref.cell(17, 1).fill = PatternFill("solid", fgColor=CALC_FILL)
ref.cell(18, 1).fill = PatternFill("solid", fgColor=HELP_FILL)

ref["A21"] = "CARA PAKAI"
ref["A21"].font = LBL
STEPS = [
    "1. Isi Daftar Affiliate. Kolom wajib: Nama, Platform, Followers, Sales per Bulan.",
    "2. Tier, komisi, GMV bersih, dan semua metrik efisiensi terisi sendiri.",
    "3. Catat GMV tiap bulan di sheet Rekap Bulanan untuk melihat trennya.",
    "4. Dashboard terisi otomatis. Tidak ada yang perlu diketik di sana.",
    "5. Hapus baris contoh (baris 6 sampai 10 di Daftar Affiliate) sebelum dipakai.",
]
for i, s in enumerate(STEPS):
    ref.cell(22 + i, 1, s).font = Font(name=F, size=10, color=INK)

ref["A29"] = "KENAPA ADA KOLOM GMV PER 1.000 FOLLOWERS"
ref["A29"].font = LBL
ref["A30"] = ("Followers besar tidak berarti penjualan besar. Kolom itu menunjukkan "
              "berapa rupiah yang dihasilkan tiap 1.000 followers,")
ref["A31"] = ("jadi affiliate kecil yang menjual efektif bisa terlihat, dan itu "
              "tidak akan kelihatan kalau daftarnya cuma diurutkan dari followers.")
for r in (30, 31):
    ref.cell(r, 1).font = Font(name=F, size=10, color=INK)

for col, w in zip("ABCDEFG", [26, 46, 4, 16, 4, 14, 4]):
    ref.column_dimensions[col].width = w

# ──────────────────────────── DAFTAR AFFILIATE ──────────────────────────────
da = wb.create_sheet("Daftar Affiliate")
da["A1"] = "DAFTAR AFFILIATE"
da["A1"].font = TITLE
da["A2"] = "Isi kolom krem saja. Kolom abu terisi sendiri."
da["A2"].font = SUB

COLS = [
    ("No",                        6,  "calc"),
    ("Nama Affiliate",           22,  "in"),
    ("Username / Handle",        20,  "in"),
    ("Platform",                 14,  "in"),
    ("Link Profil",              26,  "in"),
    ("Followers",                12,  "in"),
    ("Tier",                     10,  "calc"),
    ("Rata-rata Views/Konten",   16,  "in"),
    ("Konten/Bulan",             12,  "in"),
    ("Sales per Bulan (GMV)",    18,  "in"),
    ("Order/Bulan",              12,  "in"),
    ("Komisi %",                 10,  "in"),
    ("Biaya Komisi",             15,  "calc"),
    ("GMV Bersih",               15,  "calc"),
    ("GMV per 1.000 Followers",  18,  "calc"),
    ("Nilai Order Rata-rata",    17,  "calc"),
    ("GMV per 1.000 Views",      17,  "calc"),
    ("Status",                   11,  "in"),
    ("Tanggal Gabung",           14,  "in"),
    ("PIC",                      12,  "in"),
    ("Catatan",                  30,  "in"),
    ("Peringkat GMV",            13,  "help"),
    ("Peringkat Efisiensi",      15,  "help"),
]
for i, (h, w, kind) in enumerate(COLS, start=1):
    c = da.cell(4, i, h)
    c.font = H_FONT
    c.fill = PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
    da.column_dimensions[get_column_letter(i)].width = w
da.row_dimensions[4].height = 34

CONTOH = [
    ("Rizky Pratama",  "@rizkyfits",     "TikTok",      "tiktok.com/@rizkyfits",
     8500,     42000, 12,  6_800_000,  34, 0.12, "Aktif",  "2026-03-14", "Nuri",
     "Konversi tinggi walau followers kecil"),
    ("Bagas Adi",      "@bagasdaily",    "TikTok Shop", "tiktok.com/@bagasdaily",
     46000,   118000, 20, 21_400_000, 102, 0.10, "Aktif",  "2026-02-02", "Nuri", ""),
    ("Dimas Nugroho",  "@dimasnugroho",  "Instagram",   "instagram.com/dimasnugroho",
     132000,   31000,  8,  4_900_000,  22, 0.15, "Pantau", "2026-01-20", "Nuri",
     "Followers besar, penjualan kecil"),
    ("Aditya Wijaya",  "@adityawjy",     "TikTok",      "tiktok.com/@adityawjy",
     610000,  240000, 10, 33_500_000, 158, 0.10, "Aktif",  "2025-11-08", "Nuri", ""),
    ("Fajar Ramadhan", "@fajarrmdhn",    "Shopee Live", "shopee.co.id/fajarrmdhn",
     3200,     11000, 24,  2_150_000,  13, 0.12, "Baru",   "2026-06-30", "Nuri", ""),
]

for r in range(FIRST, LAST + 1):
    ex = CONTOH[r - FIRST] if r - FIRST < len(CONTOH) else None

    da.cell(r, 1, f'=IF($B{r}="","",ROW()-{FIRST - 1})')
    if ex:
        for col, val in zip((2, 3, 4, 5, 6), ex[:5]):
            da.cell(r, col, val)
        da.cell(r, 8, ex[5]); da.cell(r, 9, ex[6]); da.cell(r, 10, ex[7])
        da.cell(r, 11, ex[8]); da.cell(r, 12, ex[9]); da.cell(r, 18, ex[10])
        da.cell(r, 19, ex[11]); da.cell(r, 20, ex[12]); da.cell(r, 21, ex[13])

    # LOOKUP against the Referensi thresholds, not a nested IF, so the bands stay
    # editable in cells rather than buried in every one of a hundred formulas.
    da.cell(r, 7, f'=IF($F{r}="","",LOOKUP($F{r},Referensi!$A$6:$A$10,Referensi!$B$6:$B$10))')
    da.cell(r, 13, f'=IF(OR($J{r}="",$L{r}=""),"",$J{r}*$L{r})')
    da.cell(r, 14, f'=IF($J{r}="","",$J{r}-$M{r})')
    da.cell(r, 15, f'=IFERROR(IF(OR($J{r}="",$F{r}=""),"",$J{r}/$F{r}*1000),"")')
    da.cell(r, 16, f'=IFERROR(IF(OR($J{r}="",$K{r}=""),"",$J{r}/$K{r}),"")')
    da.cell(r, 17, f'=IFERROR(IF(OR($J{r}="",$H{r}="",$I{r}=""),"",$J{r}/($H{r}*$I{r})*1000),"")')
    # RANK plus a running COUNTIF makes every rank unique, so two affiliates on the
    # same GMV do not both answer the Dashboard's MATCH and hide one of themselves.
    da.cell(r, 22, f'=IF($J{r}="","",RANK($J{r},$J${FIRST}:$J${LAST},0)'
                   f'+COUNTIF($J${FIRST}:$J{r},$J{r})-1)')
    da.cell(r, 23, f'=IF($O{r}="","",RANK($O{r},$O${FIRST}:$O${LAST},0)'
                   f'+COUNTIF($O${FIRST}:$O{r},$O{r})-1)')

    for i, (_, _, kind) in enumerate(COLS, start=1):
        c = da.cell(r, i)
        c.border = BOX
        if kind == "in":
            c.font, c.fill = IN_FONT, PatternFill("solid", fgColor=INPUT_FILL)
        elif kind == "calc":
            c.font, c.fill = CA_FONT, PatternFill("solid", fgColor=CALC_FILL)
        else:
            c.font, c.fill = CA_FONT, PatternFill("solid", fgColor=HELP_FILL)

    for col in (6, 8, 9, 11):
        da.cell(r, col).number_format = NUM
    for col in (10, 13, 14, 15, 16, 17):
        da.cell(r, col).number_format = RP
    da.cell(r, 12).number_format = PCT

TOT = LAST + 2
da.cell(TOT, 2, "TOTAL").font = LBL
for col, letter in ((6, "F"), (10, "J"), (11, "K"), (13, "M"), (14, "N")):
    da.cell(TOT, col, f'=SUM({letter}{FIRST}:{letter}{LAST})')
    da.cell(TOT, col).font = Font(name=F, size=10, bold=True, color=INK)
    da.cell(TOT, col).number_format = NUM if col in (6, 11) else RP
    da.cell(TOT, col).fill = PatternFill("solid", fgColor=STEEL)
    da.cell(TOT, col).border = BOX
da.cell(TOT, 2).fill = PatternFill("solid", fgColor=STEEL)
da.cell(TOT, 2).border = BOX

da.cell(TOT + 2, 2, "Baris 6 sampai 10 adalah CONTOH. Hapus isinya sebelum dipakai; "
                    "rumusnya akan tetap jalan.").font = NOTE

# Reject anything off the list rather than merely offering one. The Dashboard
# totals by exact text, so a typed "Tiktok" would vanish from the platform
# breakdown and quietly stop the contribution column summing to 100 percent.
dv_p = DataValidation(
    type="list", formula1="=Referensi!$D$5:$D$9", allow_blank=True,
    showErrorMessage=True, errorTitle="Platform tidak dikenal",
    error="Pilih dari daftar. Ketikan bebas membuat baris ini hilang dari Dashboard.")
dv_s = DataValidation(
    type="list", formula1="=Referensi!$F$5:$F$8", allow_blank=True,
    showErrorMessage=True, errorTitle="Status tidak dikenal",
    error="Pilih dari daftar: Aktif, Baru, Pantau, atau Berhenti.")
da.add_data_validation(dv_p); da.add_data_validation(dv_s)
dv_p.add(f"D{FIRST}:D{LAST}")
dv_s.add(f"R{FIRST}:R{LAST}")

da.conditional_formatting.add(
    f"O{FIRST}:O{LAST}",
    ColorScaleRule(start_type="min", start_color="F8CBAD",
                   mid_type="percentile", mid_value=50, mid_color="FFE699",
                   end_type="max", end_color="C6E0B4"))
da.conditional_formatting.add(
    f"R{FIRST}:R{LAST}",
    CellIsRule(operator="equal", formula=['"Berhenti"'],
               font=Font(name=F, size=10, color="9C0006"),
               fill=PatternFill("solid", fgColor="FFC7CE")))
da.conditional_formatting.add(
    f"R{FIRST}:R{LAST}",
    CellIsRule(operator="equal", formula=['"Aktif"'],
               font=Font(name=F, size=10, color="006100"),
               fill=PatternFill("solid", fgColor="C6EFCE")))

da.freeze_panes = "C5"
da.auto_filter.ref = f"A4:W{LAST}"

# ───────────────────────────── REKAP BULANAN ────────────────────────────────
rb = wb.create_sheet("Rekap Bulanan")
rb["A1"] = "REKAP GMV BULANAN"
rb["A1"].font = TITLE
rb["A2"] = "Namanya ikut Daftar Affiliate. Isi GMV tiap bulan di kolom krem."
rb["A2"].font = SUB

MONTHS = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun",
          "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
heads = ["No", "Nama Affiliate"] + MONTHS + ["Total Setahun", "Rata-rata per Bulan Aktif",
                                             "Bulan Tertinggi"]
for i, h in enumerate(heads, start=1):
    c = rb.cell(4, i, h)
    c.font = H_FONT
    c.fill = PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
    rb.column_dimensions[get_column_letter(i)].width = {1: 6, 2: 24}.get(i, 13)
rb.row_dimensions[4].height = 34

for r in range(FIRST, LAST + 1):
    rb.cell(r, 1, f'=IF($B{r}="","",ROW()-{FIRST - 1})')
    rb.cell(r, 2, f"=IF('Daftar Affiliate'!B{r}=\"\",\"\",'Daftar Affiliate'!B{r})")
    rb.cell(r, 15, f'=IF($B{r}="","",SUM($C{r}:$N{r}))')
    # average over months actually worked, not over twelve. Dividing by 12 punishes
    # anyone who joined in August and makes the column useless for comparison.
    rb.cell(r, 16, f'=IFERROR(IF($B{r}="","",AVERAGEIF($C{r}:$N{r},">0")),"")')
    rb.cell(r, 17, f'=IFERROR(IF(SUM($C{r}:$N{r})=0,"",'
                   f'INDEX($C$4:$N$4,MATCH(MAX($C{r}:$N{r}),$C{r}:$N{r},0))),"")')
    for i in range(1, 18):
        c = rb.cell(r, i)
        c.border = BOX
        if 3 <= i <= 14:
            c.font, c.fill = IN_FONT, PatternFill("solid", fgColor=INPUT_FILL)
            c.number_format = RP
        else:
            c.font, c.fill = CA_FONT, PatternFill("solid", fgColor=CALC_FILL)
            if i in (15, 16):
                c.number_format = RP

rb.cell(TOT, 2, "TOTAL").font = LBL
rb.cell(TOT, 2).fill = PatternFill("solid", fgColor=STEEL)
rb.cell(TOT, 2).border = BOX
for i in range(3, 16):
    L = get_column_letter(i)
    c = rb.cell(TOT, i, f'=SUM({L}{FIRST}:{L}{LAST})')
    c.font = Font(name=F, size=10, bold=True, color=INK)
    c.number_format = RP
    c.fill = PatternFill("solid", fgColor=STEEL)
    c.border = BOX
rb.freeze_panes = "C5"

# ─────────────────────────────── DASHBOARD ──────────────────────────────────
db = wb.create_sheet("Dashboard", 0)
db["A1"] = "TONI BLACK — AFFILIATE DASHBOARD"
db["A1"].font = TITLE
db["A2"] = "Seluruh angka di sheet ini rumus. Tidak ada yang perlu diketik."
db["A2"].font = SUB

D = f"'Daftar Affiliate'!"
RNG = lambda c: f"{D}${c}${FIRST}:${c}${LAST}"


def block(row, title, rows, widths=None):
    db.cell(row, 1, title).font = LBL
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            c = db.cell(row + 1 + i, 1 + j, v)
            c.border = BOX
            if i == 0:
                c.font, c.fill = H_FONT, PatternFill("solid", fgColor=INK)
                c.alignment = Alignment(horizontal="center", wrap_text=True)
            else:
                c.font = CA_FONT
                c.fill = PatternFill("solid", fgColor=CALC_FILL)
    return row + 1 + len(rows) + 1


db.cell(4, 1, "RINGKASAN").font = LBL
SUMMARY = [
    ("Jumlah Affiliate",            f'=COUNTA({RNG("B")})', NUM),
    ("Affiliate Aktif",             f'=COUNTIF({RNG("R")},"Aktif")', NUM),
    ("Total Followers",             f'=SUM({RNG("F")})', NUM),
    ("Total GMV per Bulan",         f'=SUM({RNG("J")})', RP),
    ("Total Biaya Komisi",          f'=SUM({RNG("M")})', RP),
    ("GMV Bersih per Bulan",        f'=SUM({RNG("N")})', RP),
    ("Rata-rata GMV per Affiliate", f'=IFERROR(SUM({RNG("J")})/COUNTA({RNG("B")}),0)', RP),
    ("GMV per 1.000 Followers (keseluruhan)",
     f'=IFERROR(SUM({RNG("J")})/SUM({RNG("F")})*1000,0)', RP),
    ("Komisi Efektif",              f'=IFERROR(SUM({RNG("M")})/SUM({RNG("J")}),0)', PCT),
]
for i, (label, formula, fmt) in enumerate(SUMMARY):
    r = 5 + i
    a = db.cell(r, 1, label); a.font = Font(name=F, size=10, color=INK); a.border = BOX
    a.fill = PatternFill("solid", fgColor=CALC_FILL)
    b = db.cell(r, 2, formula); b.font = Font(name=F, size=10, bold=True, color=INK)
    b.number_format = fmt; b.border = BOX
    b.fill = PatternFill("solid", fgColor=CALC_FILL)

r0 = 16
db.cell(r0, 1, "PER TIER").font = LBL
for j, h in enumerate(["Tier", "Jumlah", "Total Followers", "Total GMV",
                       "GMV per 1.000 Followers", "Kontribusi GMV"]):
    c = db.cell(r0 + 1, 1 + j, h)
    c.font, c.fill = H_FONT, PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
for i, (_, name) in enumerate(TIERS):
    r = r0 + 2 + i
    db.cell(r, 1, name)
    db.cell(r, 2, f'=COUNTIF({RNG("G")},$A{r})')
    db.cell(r, 3, f'=SUMIF({RNG("G")},$A{r},{RNG("F")})')
    db.cell(r, 4, f'=SUMIF({RNG("G")},$A{r},{RNG("J")})')
    db.cell(r, 5, f'=IFERROR($D{r}/$C{r}*1000,0)')
    db.cell(r, 6, f'=IFERROR($D{r}/SUM({RNG("J")}),0)')
    for j, fmt in enumerate([None, NUM, NUM, RP, RP, PCT]):
        c = db.cell(r, 1 + j)
        c.font = CA_FONT; c.border = BOX
        c.fill = PatternFill("solid", fgColor=CALC_FILL)
        if fmt:
            c.number_format = fmt

r1 = r0 + 9
db.cell(r1, 1, "PER PLATFORM").font = LBL
for j, h in enumerate(["Platform", "Jumlah", "Total GMV", "Kontribusi GMV"]):
    c = db.cell(r1 + 1, 1 + j, h)
    c.font, c.fill = H_FONT, PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
for i, p in enumerate(PLATFORMS):
    r = r1 + 2 + i
    db.cell(r, 1, p)
    db.cell(r, 2, f'=COUNTIF({RNG("D")},$A{r})')
    db.cell(r, 3, f'=SUMIF({RNG("D")},$A{r},{RNG("J")})')
    db.cell(r, 4, f'=IFERROR($C{r}/SUM({RNG("J")}),0)')
    for j, fmt in enumerate([None, NUM, RP, PCT]):
        c = db.cell(r, 1 + j)
        c.font = CA_FONT; c.border = BOX
        c.fill = PatternFill("solid", fgColor=CALC_FILL)
        if fmt:
            c.number_format = fmt

# Two leaderboards side by side. The second is the one that changes decisions:
# ranked by GMV per 1.000 followers, it surfaces the small accounts that sell.
def leaderboard(top_row, col0, title, rank_col, note):
    db.cell(top_row, col0, title).font = LBL
    db.cell(top_row + 1, col0, note).font = NOTE
    for j, h in enumerate(["#", "Nama Affiliate", "Tier", "GMV per Bulan",
                           "GMV per 1.000 Followers"]):
        c = db.cell(top_row + 2, col0 + j, h)
        c.font, c.fill = H_FONT, PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
    for k in range(1, 6):
        r = top_row + 2 + k
        m = (f'MATCH({k},{RNG(rank_col)},0)')
        db.cell(r, col0, k)
        db.cell(r, col0 + 1, f'=IFERROR(INDEX({RNG("B")},{m}),"")')
        db.cell(r, col0 + 2, f'=IFERROR(INDEX({RNG("G")},{m}),"")')
        db.cell(r, col0 + 3, f'=IFERROR(INDEX({RNG("J")},{m}),"")')
        db.cell(r, col0 + 4, f'=IFERROR(INDEX({RNG("O")},{m}),"")')
        for j, fmt in enumerate([None, None, None, RP, RP]):
            c = db.cell(r, col0 + j)
            c.font = CA_FONT; c.border = BOX
            c.fill = PatternFill("solid", fgColor=CALC_FILL)
            if fmt:
                c.number_format = fmt


r2 = r1 + 8
leaderboard(r2, 1, "5 TERATAS — GMV PER BULAN", "V",
            "Siapa yang menghasilkan paling banyak.")
leaderboard(r2 + 9, 1, "5 TERATAS — GMV PER 1.000 FOLLOWERS", "W",
            "Siapa yang menjual paling efektif untuk ukurannya. Ini yang biasanya "
            "berbeda dari daftar di atas.")

for col, w in zip("ABCDEF", [34, 16, 18, 20, 24, 18]):
    db.column_dimensions[col].width = w

db.cell(r2 + 19, 1, "Semua angka menarik dari sheet Daftar Affiliate baris "
                    f"{FIRST} sampai {LAST}.").font = NOTE

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("written:", OUT)
