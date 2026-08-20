# -*- coding: utf-8 -*-
"""Membangun file Affiliate Dashboard (.xlsx)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule
from openpyxl.comments import Comment

import os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "Affiliate_Dashboard.xlsx")

FONT = "Arial"
NAVY = "1F3864"
BLUE_HDR = "2E5C8A"
LIGHT = "DCE6F1"
BAND = "F2F6FB"
GREY = "E8E8E8"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"

ROW_START = 6           # baris data pertama pada sheet Data Affiliate
ROW_END = 205           # baris data terakhir (200 baris siap isi)
DATA = "'Data Affiliate'"

FMT_RP = '"Rp" #,##0;("Rp" #,##0);"-"'
FMT_NUM = '#,##0;-#,##0;"-"'
FMT_PCT = '0.0%'

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------------------------------------------------------------- helpers
def style_title(ws, cell, text, span):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=16, bold=True, color=WHITE)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(span)
    for row in ws[span]:
        for c in row:
            c.fill = PatternFill("solid", fgColor=NAVY)

def section(ws, row, col, text, span_to):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span_to)
    for cc in range(col, span_to + 1):
        ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=BLUE_HDR)
    ws.row_dimensions[row].height = 20

def header_cells(ws, row, col_start, labels, widths=None):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=label)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box
    ws.row_dimensions[row].height = 30

# ================================================================ REFERENSI
ref = wb.active
ref.title = "Referensi"
ref.sheet_properties.tabColor = "7F7F7F"
ref.sheet_view.showGridLines = False

style_title(ref, "A1", "REFERENSI DAN PENGATURAN", "A1:F1")
ref.row_dimensions[1].height = 28

section(ref, 3, 1, "Pilihan Channel Komunikasi", 1)
CHANNELS = ["WhatsApp", "Instagram DM", "TikTok DM", "Email",
            "Telepon", "Shopee Chat", "Tokopedia Chat", "Offline / Event"]
for i, ch in enumerate(CHANNELS):
    c = ref.cell(row=4 + i, column=1, value=ch)
    c.font = Font(name=FONT, size=10, color="0000FF")
    c.border = box
    c.fill = PatternFill("solid", fgColor=WHITE)
CH_FIRST, CH_LAST = 4, 4 + len(CHANNELS) - 1

section(ref, 3, 3, "Segmentasi Jumlah Followers", 6)
header_cells(ref, 4, 3, ["Tier", "Keterangan Rentang", "Batas Bawah", "Batas Atas"])
TIERS = [
    ("Nano",  "Di bawah 10.000",            0,       9999),
    ("Micro", "10.000 sampai 49.999",       10000,   49999),
    ("Mid",   "50.000 sampai 249.999",      50000,   249999),
    ("Macro", "250.000 sampai 999.999",     250000,  999999),
    ("Mega",  "1.000.000 ke atas",          1000000, 999999999),
]
for i, (tier, ket, lo, hi) in enumerate(TIERS):
    r = 5 + i
    for j, val in enumerate((tier, ket, lo, hi)):
        c = ref.cell(row=r, column=3 + j, value=val)
        c.font = Font(name=FONT, size=10, bold=(j == 0),
                      color="0000FF" if j >= 2 else "000000")
        c.border = box
        c.alignment = Alignment(horizontal="center" if j >= 2 else "left")
        if j >= 2:
            c.number_format = FMT_NUM
TIER_FIRST, TIER_LAST = 5, 5 + len(TIERS) - 1

ref["C11"] = "Catatan: angka biru adalah nilai yang boleh diubah manual. Mengubah batas tier di sini otomatis memperbarui perhitungan di sheet Dashboard."
ref["C11"].font = Font(name=FONT, size=9, italic=True, color="595959")
ref.merge_cells("C11:F11")

ref["A13"] = "Catatan: menambah atau mengganti nama channel di kolom A akan otomatis terbaca oleh dropdown dan oleh tabel breakdown di sheet Dashboard."
ref["A13"].font = Font(name=FONT, size=9, italic=True, color="595959")
ref.merge_cells("A13:F13")

for col, w in zip("ABCDEF", (22, 4, 16, 26, 16, 16)):
    ref.column_dimensions[col].width = w

# ================================================================ DATA
ws = wb.create_sheet("Data Affiliate")
ws.sheet_properties.tabColor = "2E5C8A"
ws.sheet_view.showGridLines = False

style_title(ws, "A1", "DATABASE AFFILIATE", "A1:H1")
ws.row_dimensions[1].height = 30

ws["A2"] = "Isi tabel di bawah ini. Semua ringkasan di sheet Dashboard terhitung otomatis dari data ini."
ws["A2"].font = Font(name=FONT, size=10, italic=True, color="595959")
ws.merge_cells("A2:H2")

ws["A3"] = "Cara pakai: kolom B sampai H diisi manual mulai baris 6. Kolom A (No) berisi rumus otomatis, jangan diketik. Baris 6 adalah baris contoh, silakan timpa dengan data asli. Kapasitas tabel 200 baris (baris 6 sampai 205)."
ws["A3"].font = Font(name=FONT, size=9, color="806000")
ws["A3"].fill = PatternFill("solid", fgColor=YELLOW)
ws["A3"].alignment = Alignment(vertical="center", wrap_text=True, indent=1)
ws.merge_cells("A3:H3")
ws.row_dimensions[3].height = 28
ws.row_dimensions[4].height = 6

HEADERS = ["No", "Nama", "Nama Akun", "Nomor Whatsapp", "Dihubungi Lewat Apa?",
           "GMV Histori (Rp)", "Product yang Dikirim", "Jumlah Followers"]
header_cells(ws, 5, 1, HEADERS)

ws["D5"].comment = Comment(
    "Ketik nomor lengkap dengan angka 0 di depan, contoh 081234567890. "
    "Kolom ini sudah diformat sebagai teks sehingga angka 0 di depan tidak hilang.", "Panduan")
ws["F5"].comment = Comment(
    "Isi angka polos tanpa titik dan tanpa tulisan Rp, contoh 45750000. "
    "Format Rupiah muncul otomatis.", "Panduan")
ws["H5"].comment = Comment(
    "Isi angka polos tanpa titik, contoh 128000. Dipakai untuk segmentasi tier di sheet Dashboard.", "Panduan")

CONTOH = ["Rina Kartika", "@rinakartika.id", "081234567890", "Instagram DM",
          45750000, "Tote Bag Series A (2 pcs)", 128000]

for r in range(ROW_START, ROW_END + 1):
    is_contoh = (r == ROW_START)
    fill = PatternFill("solid", fgColor=YELLOW if is_contoh else WHITE)

    a = ws.cell(row=r, column=1, value='=IF($B{0}<>"",ROW()-{1},"")'.format(r, ROW_START - 1))
    a.font = Font(name=FONT, size=10, bold=True, color="595959")
    a.fill = PatternFill("solid", fgColor=GREY)
    a.alignment = Alignment(horizontal="center", vertical="center")
    a.border = box

    for j in range(2, 9):
        c = ws.cell(row=r, column=j)
        if is_contoh:
            c.value = CONTOH[j - 2]
        c.font = Font(name=FONT, size=10)
        c.fill = fill
        c.border = box
        if j == 4:
            c.number_format = "@"
            c.alignment = Alignment(horizontal="left", vertical="center")
        elif j == 6:
            c.number_format = FMT_RP
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif j == 8:
            c.number_format = FMT_NUM
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(j == 7))
    ws.row_dimensions[r].height = 18

dv_ch = DataValidation(type="list", formula1="=Referensi!$A${0}:$A${1}".format(CH_FIRST, CH_LAST),
                       allow_blank=True, showDropDown=False)
dv_ch.error = "Pilih salah satu channel dari daftar dropdown."
dv_ch.errorTitle = "Channel tidak dikenal"
dv_ch.prompt = "Pilih channel dari dropdown."
dv_ch.promptTitle = "Dihubungi lewat apa?"
ws.add_data_validation(dv_ch)
dv_ch.add("E{0}:E{1}".format(ROW_START, ROW_END))

dv_gmv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv_gmv.error = "GMV Histori harus berupa angka dan tidak boleh negatif."
dv_gmv.errorTitle = "Nilai tidak valid"
ws.add_data_validation(dv_gmv)
dv_gmv.add("F{0}:F{1}".format(ROW_START, ROW_END))

dv_fol = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv_fol.error = "Jumlah followers harus berupa angka bulat dan tidak boleh negatif."
dv_fol.errorTitle = "Nilai tidak valid"
ws.add_data_validation(dv_fol)
dv_fol.add("H{0}:H{1}".format(ROW_START, ROW_END))

ws.conditional_formatting.add(
    "F{0}:F{1}".format(ROW_START, ROW_END),
    DataBarRule(start_type="num", start_value=0, end_type="max", color="63BE7B", showValue=True))
ws.conditional_formatting.add(
    "H{0}:H{1}".format(ROW_START, ROW_END),
    DataBarRule(start_type="num", start_value=0, end_type="max", color="8EA9DB", showValue=True))

ws.auto_filter.ref = "A5:H{0}".format(ROW_END)
ws.freeze_panes = "B6"

for col, w in zip("ABCDEFGH", (6, 24, 22, 20, 20, 18, 34, 16)):
    ws.column_dimensions[col].width = w

# ================================================================ DASHBOARD
dsh = wb.create_sheet("Dashboard", 0)
dsh.sheet_properties.tabColor = "C00000"
dsh.sheet_view.showGridLines = False

style_title(dsh, "B2", "AFFILIATE DASHBOARD", "B2:G2")
dsh.row_dimensions[2].height = 34
dsh["B3"] = "Ringkasan otomatis dari sheet Data Affiliate. Seluruh angka di halaman ini adalah rumus, jangan diketik manual."
dsh["B3"].font = Font(name=FONT, size=10, italic=True, color="595959")
dsh.merge_cells("B3:G3")

R = "{0}!$".format(DATA)
rng = lambda col: "{0}!${1}${2}:${1}${3}".format(DATA, col, ROW_START, ROW_END)

# --- KPI ---------------------------------------------------------------
section(dsh, 5, 2, "RINGKASAN UTAMA", 7)
header_cells(dsh, 6, 2, ["Metrik", "Nilai", "Keterangan"])
for cc in range(5, 8):
    hc = dsh.cell(row=6, column=cc)
    hc.fill = PatternFill("solid", fgColor=BLUE_HDR)
    hc.border = box
dsh.merge_cells("D6:G6")

KPI = [
    ("Total Affiliate Terdata", "=COUNTA({0})".format(rng("B")), FMT_NUM,
     "Jumlah baris yang sudah terisi nama."),
    ("Total GMV Histori", "=SUM({0})".format(rng("F")), FMT_RP,
     "Penjumlahan seluruh GMV histori affiliate."),
    ("Rata-rata GMV per Affiliate", "=IF($C$7=0,0,$C$8/$C$7)", FMT_RP,
     "Total GMV dibagi jumlah affiliate terdata."),
    ("GMV Histori Tertinggi", "=MAX({0})".format(rng("F")), FMT_RP,
     "Nilai GMV terbesar dalam database."),
    ("Affiliate dengan GMV Tertinggi",
     '=IFERROR(INDEX({0},MATCH(MAX({1}),{1},0)),"-")'.format(rng("B"), rng("F")), "General",
     "Nama pemilik GMV tertinggi."),
    ("Total Followers", "=SUM({0})".format(rng("H")), FMT_NUM,
     "Total jangkauan followers seluruh affiliate."),
    ("Rata-rata Followers", "=IF($C$7=0,0,$C$12/$C$7)", FMT_NUM,
     "Total followers dibagi jumlah affiliate terdata."),
    ("Sudah Dikirim Produk", '=SUMPRODUCT(({0}<>"")*1)'.format(rng("G")), FMT_NUM,
     "Baris yang kolom Product yang Dikirim sudah terisi."),
    ("Belum Dikirim Produk", "=MAX(0,$C$7-$C$14)", FMT_NUM,
     "Affiliate terdata yang belum menerima kiriman produk."),
    ("Belum Dihubungi", '=SUMPRODUCT(({0}<>"")*({1}=""))'.format(rng("B"), rng("E")), FMT_NUM,
     "Nama sudah ada tetapi channel komunikasi masih kosong."),
]
for i, (label, formula, fmt, note) in enumerate(KPI):
    r = 7 + i
    c1 = dsh.cell(row=r, column=2, value=label)
    c1.font = Font(name=FONT, size=10, bold=True)
    c1.alignment = Alignment(vertical="center", indent=1)
    c2 = dsh.cell(row=r, column=3, value=formula)
    c2.font = Font(name=FONT, size=11, bold=True, color="006100")
    c2.number_format = fmt
    c2.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c3 = dsh.cell(row=r, column=4, value=note)
    c3.font = Font(name=FONT, size=9, color="595959")
    c3.alignment = Alignment(vertical="center", indent=1)
    dsh.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    band = BAND if i % 2 == 0 else WHITE
    for cc in range(2, 8):
        cell = dsh.cell(row=r, column=cc)
        cell.border = box
        cell.fill = PatternFill("solid", fgColor=band)
    dsh.row_dimensions[r].height = 19

KPI_END = 7 + len(KPI) - 1   # 16

# --- BREAKDOWN CHANNEL -------------------------------------------------
CH_TOP = KPI_END + 2                 # 18 judul
section(dsh, CH_TOP, 2, "BREAKDOWN CHANNEL KOMUNIKASI", 7)
header_cells(dsh, CH_TOP + 1, 2,
             ["Channel", "Jumlah Affiliate", "Total GMV Histori",
              "Rata-rata GMV", "Total Followers", "% Kontribusi GMV"])
CH_R1 = CH_TOP + 2
CH_R2 = CH_R1 + len(CHANNELS) - 1
CH_TOT = CH_R2 + 1

for i in range(len(CHANNELS)):
    r = CH_R1 + i
    dsh.cell(row=r, column=2, value="=Referensi!$A${0}".format(CH_FIRST + i)).font = \
        Font(name=FONT, size=10, color="008000")
    dsh.cell(row=r, column=3, value="=COUNTIFS({0},$B{1})".format(rng("E"), r))
    dsh.cell(row=r, column=4, value="=SUMIFS({0},{1},$B{2})".format(rng("F"), rng("E"), r))
    dsh.cell(row=r, column=5, value="=IF($C{0}=0,0,$D{0}/$C{0})".format(r))
    dsh.cell(row=r, column=6, value="=SUMIFS({0},{1},$B{2})".format(rng("H"), rng("E"), r))
    dsh.cell(row=r, column=7, value="=IF($D${0}=0,0,$D{1}/$D${0})".format(CH_TOT, r))

dsh.cell(row=CH_TOT, column=2, value="TOTAL / TERISI")
dsh.cell(row=CH_TOT, column=3, value="=SUM($C${0}:$C${1})".format(CH_R1, CH_R2))
dsh.cell(row=CH_TOT, column=4, value="=SUM($D${0}:$D${1})".format(CH_R1, CH_R2))
dsh.cell(row=CH_TOT, column=5, value="=IF($C${0}=0,0,$D${0}/$C${0})".format(CH_TOT))
dsh.cell(row=CH_TOT, column=6, value="=SUM($F${0}:$F${1})".format(CH_R1, CH_R2))
dsh.cell(row=CH_TOT, column=7, value="=IF($D${0}=0,0,SUM($G${1}:$G${2}))".format(CH_TOT, CH_R1, CH_R2))

for r in range(CH_R1, CH_TOT + 1):
    total = (r == CH_TOT)
    for cc in range(2, 8):
        cell = dsh.cell(row=r, column=cc)
        cell.border = box
        cell.fill = PatternFill("solid", fgColor=LIGHT if total else
                                (BAND if (r - CH_R1) % 2 == 0 else WHITE))
        if cc == 2:
            cell.font = Font(name=FONT, size=10, bold=total,
                             color="000000" if total else "008000")
            cell.alignment = Alignment(vertical="center", indent=1)
        else:
            cell.font = Font(name=FONT, size=10, bold=total)
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.number_format = {3: FMT_NUM, 4: FMT_RP, 5: FMT_RP,
                                  6: FMT_NUM, 7: FMT_PCT}[cc]
    dsh.row_dimensions[r].height = 18

dsh.cell(row=CH_TOT + 1, column=2,
         value="Catatan: baris yang kolom channelnya masih kosong tidak masuk hitungan tabel ini. Lihat metrik Belum Dihubungi di ringkasan utama.")
dsh.cell(row=CH_TOT + 1, column=2).font = Font(name=FONT, size=9, italic=True, color="595959")
dsh.merge_cells(start_row=CH_TOT + 1, start_column=2, end_row=CH_TOT + 1, end_column=7)

# --- SEGMENTASI FOLLOWERS ---------------------------------------------
TR_TOP = CH_TOT + 3
section(dsh, TR_TOP, 2, "SEGMENTASI JUMLAH FOLLOWERS", 7)
header_cells(dsh, TR_TOP + 1, 2,
             ["Tier", "Rentang Followers", "Jumlah Affiliate",
              "Total GMV Histori", "Rata-rata GMV", "% Kontribusi GMV"])
TR_R1 = TR_TOP + 2
TR_R2 = TR_R1 + len(TIERS) - 1
TR_TOT = TR_R2 + 1

for i in range(len(TIERS)):
    r = TR_R1 + i
    lo = "Referensi!$E${0}".format(TIER_FIRST + i)
    hi = "Referensi!$F${0}".format(TIER_FIRST + i)
    dsh.cell(row=r, column=2, value="=Referensi!$C${0}".format(TIER_FIRST + i))
    dsh.cell(row=r, column=3, value="=Referensi!$D${0}".format(TIER_FIRST + i))
    dsh.cell(row=r, column=4,
             value='=COUNTIFS({0},">="&{1},{0},"<="&{2})'.format(rng("H"), lo, hi))
    dsh.cell(row=r, column=5,
             value='=SUMIFS({0},{1},">="&{2},{1},"<="&{3})'.format(rng("F"), rng("H"), lo, hi))
    dsh.cell(row=r, column=6, value="=IF($D{0}=0,0,$E{0}/$D{0})".format(r))
    dsh.cell(row=r, column=7, value="=IF($E${0}=0,0,$E{1}/$E${0})".format(TR_TOT, r))

dsh.cell(row=TR_TOT, column=2, value="TOTAL")
dsh.cell(row=TR_TOT, column=3, value="Seluruh tier")
dsh.cell(row=TR_TOT, column=4, value="=SUM($D${0}:$D${1})".format(TR_R1, TR_R2))
dsh.cell(row=TR_TOT, column=5, value="=SUM($E${0}:$E${1})".format(TR_R1, TR_R2))
dsh.cell(row=TR_TOT, column=6, value="=IF($D${0}=0,0,$E${0}/$D${0})".format(TR_TOT))
dsh.cell(row=TR_TOT, column=7, value="=IF($E${0}=0,0,SUM($G${1}:$G${2}))".format(TR_TOT, TR_R1, TR_R2))

for r in range(TR_R1, TR_TOT + 1):
    total = (r == TR_TOT)
    for cc in range(2, 8):
        cell = dsh.cell(row=r, column=cc)
        cell.border = box
        cell.fill = PatternFill("solid", fgColor=LIGHT if total else
                                (BAND if (r - TR_R1) % 2 == 0 else WHITE))
        if cc in (2, 3):
            cell.font = Font(name=FONT, size=10, bold=(cc == 2 or total),
                             color="000000" if total else ("008000" if not total else "000000"))
            cell.alignment = Alignment(vertical="center", indent=1)
        else:
            cell.font = Font(name=FONT, size=10, bold=total)
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.number_format = {4: FMT_NUM, 5: FMT_RP, 6: FMT_RP, 7: FMT_PCT}[cc]
    dsh.row_dimensions[r].height = 18

dsh.cell(row=TR_TOT + 1, column=2,
         value="Batas tiap tier diatur di sheet Referensi kolom E dan F, silakan ubah bila definisi tier berbeda.")
dsh.cell(row=TR_TOT + 1, column=2).font = Font(name=FONT, size=9, italic=True, color="595959")
dsh.merge_cells(start_row=TR_TOT + 1, start_column=2, end_row=TR_TOT + 1, end_column=7)

# --- LEGENDA ----------------------------------------------------------
LG = TR_TOT + 3
section(dsh, LG, 2, "LEGENDA DAN CARA PAKAI", 7)
LEGEND = [
    (YELLOW, "000000", "Sel kuning",
     "Baris contoh di sheet Data Affiliate baris 6. Timpa dengan data asli Anda."),
    (WHITE, "000000", "Sel putih",
     "Sel input manual. Isi kolom B sampai H sheet Data Affiliate mulai baris 6 sampai 205."),
    (GREY, "595959", "Sel abu-abu",
     "Kolom No pada sheet Data Affiliate. Berisi rumus penomoran otomatis, jangan diketik."),
    (BAND, "006100", "Angka hijau",
     "Hasil rumus otomatis di sheet Dashboard dan angka yang menarik data dari sheet lain."),
    (BAND, "0000FF", "Angka biru",
     "Nilai pengaturan di sheet Referensi seperti daftar channel dan batas tier followers."),
]
for i, (fill, color, label, desc) in enumerate(LEGEND):
    r = LG + 1 + i
    c1 = dsh.cell(row=r, column=2, value=label)
    c1.fill = PatternFill("solid", fgColor=fill)
    c1.font = Font(name=FONT, size=10, bold=True, color=color)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = box
    c2 = dsh.cell(row=r, column=3, value=desc)
    c2.font = Font(name=FONT, size=10)
    c2.alignment = Alignment(vertical="center", indent=1)
    dsh.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    for cc in range(3, 8):
        dsh.cell(row=r, column=cc).border = box
    dsh.row_dimensions[r].height = 19

NOTE = LG + len(LEGEND) + 2
dsh.cell(row=NOTE, column=2,
         value="Asumsi yang dipakai file ini: kapasitas tabel 200 baris (Data Affiliate baris 6 sampai 205), GMV Histori dalam Rupiah penuh tanpa pemisah ribuan saat diketik, dan tier followers memakai standar umum industri (Nano, Micro, Mid, Macro, Mega) yang batasnya bisa diubah sendiri di sheet Referensi. Daftar channel komunikasi disusun berdasarkan kanal yang lazim dipakai tim affiliate di Indonesia, bukan angka resmi dari sumber eksternal.")
dsh.cell(row=NOTE, column=2).font = Font(name=FONT, size=9, italic=True, color="595959")
dsh.cell(row=NOTE, column=2).alignment = Alignment(vertical="top", wrap_text=True, indent=1)
dsh.merge_cells(start_row=NOTE, start_column=2, end_row=NOTE + 1, end_column=7)

for col, w in zip("ABCDEFG", (3, 30, 22, 22, 20, 20, 20)):
    dsh.column_dimensions[col].width = w
dsh.column_dimensions["D"].width = 24

dsh.sheet_view.zoomScale = 100

# --- pengaturan cetak --------------------------------------------------
dsh.page_setup.orientation = "landscape"
dsh.page_setup.fitToWidth = 1
dsh.page_setup.fitToHeight = 0
dsh.sheet_properties.pageSetUpPr.fitToPage = True
dsh.print_area = "B2:G{0}".format(NOTE + 1)
dsh.page_margins.left = dsh.page_margins.right = 0.4
dsh.oddFooter.center.text = "Affiliate Dashboard - halaman &P dari &N"
dsh.oddFooter.center.size = 8

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = "5:5"
ws.page_margins.left = ws.page_margins.right = 0.4
ws.oddFooter.center.text = "Database Affiliate - halaman &P dari &N"
ws.oddFooter.center.size = 8

ref.page_setup.orientation = "landscape"
ref.page_setup.fitToWidth = 1
ref.page_setup.fitToHeight = 0
ref.sheet_properties.pageSetUpPr.fitToPage = True

wb.move_sheet("Data Affiliate", offset=-1)   # urutan: Dashboard, Data Affiliate, Referensi
wb.active = 0
wb.save(OUT)
print("saved:", OUT)
print("KPI_END", KPI_END, "CH", CH_R1, CH_R2, CH_TOT, "TR", TR_R1, TR_R2, TR_TOT, "LG", LG, "NOTE", NOTE)
