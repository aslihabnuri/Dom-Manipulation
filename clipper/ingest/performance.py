"""Parse hour-by-hour live performance exports.

Seller Center exports vary by locale, language and report type, so instead of
demanding one fixed layout we match columns by alias and parse numbers in both
Indonesian (1.234,56 / Rp / rb / jt) and English (1,234.56) conventions.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Sequence

from ..models import PerfRow

# Column aliases, matched against normalised headers. Longer/more specific
# aliases are checked first so "total penonton" wins over "penonton".
COLUMN_ALIASES: dict[str, list[str]] = {
    "time": [
        "rentang waktu", "periode waktu", "waktu mulai", "jam mulai", "start time",
        "time range", "time period", "waktu", "jam", "time", "hour", "hourly", "tanggal jam",
        "timestamp", "periode",
    ],
    "viewers": [
        "total penonton", "jumlah penonton", "penonton unik", "puncak penonton",
        "peak viewers", "total viewers", "unique viewers", "live viewers",
        "penonton", "pemirsa", "viewers", "views", "penayangan", "tayangan",
        "audience", "pv", "uv",
    ],
    "gmv": [
        "total pendapatan", "pendapatan kotor", "nilai transaksi", "gross revenue",
        "total gmv", "revenue", "pendapatan", "penjualan", "omzet", "omset",
        "gmv", "sales", "sales amount", "nilai penjualan",
    ],
    "orders": [
        "jumlah pesanan", "total pesanan", "pesanan dibuat", "orders created",
        "total orders", "produk terjual", "items sold", "unit terjual",
        "pesanan", "orders", "order", "transaksi", "qty", "kuantitas",
    ],
    "duration": ["durasi", "duration", "lama live", "live duration"],
}

_MULTIPLIERS = {
    "rb": 1_000, "ribu": 1_000, "k": 1_000,
    "jt": 1_000_000, "juta": 1_000_000, "m": 1_000_000, "mn": 1_000_000,
    "mio": 1_000_000, "b": 1_000_000_000, "miliar": 1_000_000_000, "milyar": 1_000_000_000,
}


def normalize_header(value: str) -> str:
    """Lowercase, strip accents and punctuation, collapse whitespace."""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = re.sub(r"[\(\[].*?[\)\]]", " ", text)  # drop "(Rp)", "[IDR]" etc.
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_number(value: Any) -> float:
    """Parse a numeric cell written in Indonesian or English convention.

    Handles ``Rp1.234.567``, ``1,234.56``, ``1.234,56``, ``12rb``, ``1,2 jt``,
    ``-``, ``N/A`` and percentages.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = str(value).strip()
    if not text or text.lower() in {"-", "--", "n/a", "na", "null", "none", ""}:
        return 0.0

    text = re.sub(r"(?i)\b(rp|idr|usd|\$)\b\.?", "", text)
    text = text.replace(" ", " ").strip()

    multiplier = 1.0
    m = re.search(r"(?i)([a-z]+)\s*$", text)
    if m:
        suffix = m.group(1).lower()
        if suffix in _MULTIPLIERS:
            multiplier = float(_MULTIPLIERS[suffix])
            text = text[: m.start()].strip()

    text = text.rstrip("%").strip()
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return 0.0

    has_comma, has_dot = "," in text, "." in text
    if has_comma and has_dot:
        # Whichever separator appears last is the decimal point.
        decimal_sep = "," if text.rfind(",") > text.rfind(".") else "."
        thousands_sep = "." if decimal_sep == "," else ","
        text = text.replace(thousands_sep, "").replace(decimal_sep, ".")
    elif has_comma:
        parts = text.split(",")
        # "1,234" with a 3-digit tail is thousands; "1,5" is a decimal.
        if len(parts) > 2 or (len(parts[-1]) == 3 and len(parts[0]) <= 3 and parts[0] not in ("0", "-0")):
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
    elif has_dot:
        parts = text.split(".")
        if len(parts) > 2 or (len(parts[-1]) == 3 and parts[0] not in ("0", "-0")):
            text = text.replace(".", "")

    try:
        return float(text) * multiplier
    except ValueError:
        return 0.0


def parse_time(value: Any, base_date: datetime | None = None) -> datetime | None:
    """Parse a time cell into a datetime, using ``base_date`` for date-less values."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    full_formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
    ]

    def try_full(candidate: str) -> datetime | None:
        for fmt in full_formats:
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
        return None

    # Try the whole cell first: an ISO date like "2026-08-18 09:00" contains
    # hyphens that would otherwise be mistaken for a range separator.
    parsed = try_full(text)
    if parsed:
        return parsed

    # "14:00 - 15:00", "14:00~15:00", "09:00 s/d 10:00" -> keep the start. A bare
    # hyphen only splits when spaced, so in-date hyphens survive.
    text = re.split(r"(?:\s+-\s+|\s*[–—~]\s*|\s+(?:to|s/d|sampai)\s+)", text, flags=re.I)[0].strip()

    parsed = try_full(text)
    if parsed:
        return parsed

    base = base_date or datetime(1970, 1, 1)
    base = base.replace(hour=0, minute=0, second=0, microsecond=0)

    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I %p", "%I:%M%p", "%I%p"):
        try:
            t = datetime.strptime(text.upper().replace(".", ":"), fmt)
            return base.replace(hour=t.hour, minute=t.minute, second=t.second)
        except ValueError:
            continue

    if re.fullmatch(r"\d{1,2}", text):  # bare hour, e.g. "14"
        hour = int(text)
        if 0 <= hour <= 23:
            return base.replace(hour=hour)
    return None


def match_columns(headers: Sequence[str]) -> dict[str, int]:
    """Map logical field names to column indices."""
    normalized = [normalize_header(h) for h in headers]
    mapping: dict[str, int] = {}
    taken: set[int] = set()

    for field, aliases in COLUMN_ALIASES.items():
        best: tuple[int, int] | None = None  # (alias length, column index)
        for alias in aliases:
            for idx, header in enumerate(normalized):
                if idx in taken or not header:
                    continue
                if header == alias:
                    best = (len(alias) + 100, idx)  # exact match outranks substring
                    break
                if alias in header and (best is None or len(alias) > best[0]):
                    best = (len(alias), idx)
            if best and best[0] > 100:
                break
        if best:
            mapping[field] = best[1]
            taken.add(best[1])
    return mapping


def _read_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "reading .xlsx needs openpyxl -- run: pip install openpyxl, "
                "or export the report as CSV"
            ) from exc
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")

        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = max(",;\t|", key=sample.count)
        rows = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)]

    rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    if not rows:
        raise ValueError(f"no data rows in {path}")

    # Seller Center often prefixes exports with title/filter lines; the real
    # header is the first row that matches at least two known columns.
    header_idx = 0
    for idx, row in enumerate(rows[:15]):
        cells = [str(c) if c is not None else "" for c in row]
        if len(match_columns(cells)) >= 2:
            header_idx = idx
            break

    headers = [str(c) if c is not None else "" for c in rows[header_idx]]
    return headers, rows[header_idx + 1 :]


def load_performance(
    path: str | Path,
    base_date: datetime | None = None,
    default_slot_minutes: int = 60,
) -> list[PerfRow]:
    """Load a performance export into ``PerfRow`` objects, sorted by start time."""
    p = Path(path).expanduser()
    if not p.exists():
        raise FileNotFoundError(f"performance file not found: {p}")

    headers, data_rows = _read_rows(p)
    mapping = match_columns(headers)

    if "time" not in mapping:
        raise ValueError(
            f"could not find a time/hour column in {p.name}. "
            f"Headers seen: {[h for h in headers if h][:12]}"
        )
    if "viewers" not in mapping and "gmv" not in mapping and "orders" not in mapping:
        raise ValueError(
            f"found no viewers/GMV/orders column in {p.name}. "
            f"Headers seen: {[h for h in headers if h][:12]}"
        )

    def cell(row: Sequence[Any], field: str) -> Any:
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows: list[PerfRow] = []
    for raw in data_rows:
        start = parse_time(cell(raw, "time"), base_date)
        if start is None:
            continue

        duration_min = parse_number(cell(raw, "duration")) if "duration" in mapping else 0.0
        slot = int(duration_min) if duration_min > 0 else default_slot_minutes

        extra = {
            normalize_header(h): raw[i]
            for i, h in enumerate(headers)
            if i < len(raw) and h and i not in mapping.values()
        }

        rows.append(
            PerfRow(
                start=start,
                end=start + timedelta(minutes=slot),
                viewers=parse_number(cell(raw, "viewers")),
                gmv=parse_number(cell(raw, "gmv")),
                orders=parse_number(cell(raw, "orders")),
                extra=extra,
            )
        )

    if not rows:
        raise ValueError(f"no rows with a parseable time value in {p.name}")

    rows.sort(key=lambda r: r.start)

    # Rows are often exported without a duration; infer each slot from the gap
    # to the next row so 15-minute and 60-minute reports both work.
    if "duration" not in mapping and len(rows) > 1:
        gaps = [
            (rows[i + 1].start - rows[i].start).total_seconds()
            for i in range(len(rows) - 1)
            if rows[i + 1].start > rows[i].start
        ]
        if gaps:
            gaps.sort()
            cadence = timedelta(seconds=gaps[len(gaps) // 2])  # median, ignores gaps in the export
            for i, row in enumerate(rows):
                if i + 1 < len(rows):
                    gap = rows[i + 1].start - row.start
                    slot = gap if timedelta(0) < gap <= cadence else cadence
                else:
                    # The final row has no successor, so give it the same
                    # cadence rather than leaving the 60-minute default.
                    slot = cadence
                row.end = row.start + slot

    return rows


def describe_mapping(path: str | Path) -> dict[str, str]:
    """Show which spreadsheet column was picked for each field (for `clipper doctor`)."""
    headers, _ = _read_rows(Path(path).expanduser())
    mapping = match_columns(headers)
    return {field: headers[idx] for field, idx in mapping.items()}
